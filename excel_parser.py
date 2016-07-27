from openpyxl import Workbook
from openpyxl import load_workbook

def read_words_from_file(fileName, sheetName, startRow, endRow):
	wb = load_workbook(fileName)
	ws = wb.get_sheet_by_name(sheetName)

	allwords = []
	for word in ws.iter_rows("B"+str(startRow)+":B"+str(endRow)):
		print word[0].value.strip()
		try:
			allwords.append(word[0].value.strip())
		except:
			try:
				allwords.append(word[0].value.decode("utf-8").encode("ascii", "ignore"))
			except:
				allwords.append("NaN")
	return allwords
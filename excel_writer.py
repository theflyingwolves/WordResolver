from openpyxl import Workbook
from openpyxl import load_workbook

def write_to_excel_column(filename, sheetname, startRow, endRow, data):
	wb = load_workbook(filename)
	ws = wb.get_sheet_by_name(sheetname)
	currPos = startRow
	for cellRow in ws.iter_rows("D"+str(startRow)+":D"+str(endRow)):
		cellRow[0].value = data[currPos-startRow]
		currPos+=1
	wb.save(filename)